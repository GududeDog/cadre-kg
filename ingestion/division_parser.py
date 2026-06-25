"""领导班子成员分工备案表 .xlsx 解析模块。

从 Excel 表格中逐行提取干部的姓名（编号）、分工内容、分管部门，
生成 Division 实体和对应的 Cadre 节点。
"""
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


def parse_division_table(
    xlsx_path,
    source_doc: Optional[str] = None,
) -> Tuple[List[Dict], List[Dict]]:
    """解析分工备案表，返回 (entities, relations)。

    每条数据行生成：
      - 1 个 Cadre 实体
      - 1 个 Division 实体
      - 1 条 HAS_DIVISION 关系
    """
    fp = Path(xlsx_path)
    doc_name = source_doc or fp.name

    wb = openpyxl.load_workbook(str(fp), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # ── 定位表头行 ──
    header_row = _find_header_row(ws)
    if header_row is None:
        wb.close()
        return [], []

    # ── 映射列: 姓名/职务/分工内容/分管部门 → column index ──
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = _cell_text(ws, header_row, c).replace(" ", "")
        if "姓名" in v or "姓" in v:
            col_map["cadre_col"] = c
        elif "分工" in v:
            col_map["content_col"] = c
        elif "分管" in v or "部门" in v:
            col_map["dept_col"] = c

    if "cadre_col" not in col_map or "content_col" not in col_map:
        wb.close()
        return [], []

    # ── 提取数据行 ──
    entities: List[Dict] = []
    relations: List[Dict] = []
    seq = 1

    for r in range(header_row + 1, ws.max_row + 1):
        cadre_id = _cell_text(ws, r, col_map["cadre_col"]).strip()
        if not cadre_id or cadre_id.startswith("填报") or len(cadre_id) > 10:
            continue  # 跳过空行/说明行

        content = _cell_text(ws, r, col_map["content_col"]).strip()
        department = ""
        if "dept_col" in col_map:
            department = _cell_text(ws, r, col_map["dept_col"]).strip()

        # Cadre 实体
        entities.append({
            "type": "Cadre",
            "name": cadre_id,
            "properties": {"cadre_id": cadre_id, "name": cadre_id},
        })

        # Division 实体
        div_id = f"division_{cadre_id}_{seq:02d}"
        entities.append({
            "type": "Division",
            "name": div_id,
            "properties": {
                "division_id": div_id,
                "cadre_id": cadre_id,
                "division_content": content,
                "department": department,
                "source_doc": doc_name,
            },
        })

        # 关系
        relations.append({
            "source_type": "Cadre",
            "source_name": cadre_id,
            "relation": "HAS_DIVISION",
            "target_type": "Division",
            "target_name": div_id,
            "properties": {},
        })

        seq += 1

    wb.close()
    return entities, relations


def _find_header_row(ws, max_scan: int = 10) -> Optional[int]:
    """扫描前 max_scan 行，找包含'姓名'或'分工'或'分管'的行。"""
    for r in range(1, min(ws.max_row + 1, max_scan + 1)):
        row_text = ""
        for c in range(1, ws.max_column + 1):
            v = _cell_text(ws, r, c)
            row_text += v
        # 去掉空格后匹配（表头可能是"姓 名"而非"姓名"）
        compact = row_text.replace(" ", "")
        if ("姓名" in compact or "姓" in compact) and "分工" in compact:
            return r
    return None


def _cell_text(ws, row: int, col: int) -> str:
    """读取单元格文本，合并区域取左上角的值。"""
    v = ws.cell(row, col).value
    if v is None:
        return ""
    s = str(v).strip()
    # Excel 文本格式标记：去掉前导 ' 字符（如 '001 → 001, '姓 名 → 姓 名）
    if s.startswith("'") and len(s) > 1:
        s = s[1:].strip()
    return s
